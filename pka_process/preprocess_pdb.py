import os
import shutil
import requests
import csv
import numpy as np
import pandas as pd
import re
import time
from openpyxl import Workbook, load_workbook

amino_acids_names = ['MSE', 'ALA', 'GLY', 'ILE', 'LEU', 'PRO', 'VAL',
                     'PHE', 'TRP', 'TYR', 'ASP', 'GLU', 'ARG', 'HIS',
                     'LYS', 'SER', 'THR', 'CYS', 'MET', 'ASN', 'GLN']


def cp_NMR_pdb(input_dir, output_dir):
    """
    this function will find NMR pdb files, and copy it from input_dir to output_dir.
    :param input_dir: String, path of source dir countain pdb files.
    :param output_dir: String, path of target dir to save pdb dir.
    :return: None
    """
    pdb_list = os.listdir(input_dir)
    # read pdb dir
    for pdb_name in pdb_list:
        input_pdb = os.path.join(input_dir, pdb_name)
        output_pdb = os.path.join(output_dir, pdb_name)
        # read pdb file and cp NMR pdb
        with open(input_pdb, 'r') as in_f:
            end_title = False
            line = in_f.readline()
            title_info = []
            while line:
                # print(line, end='')
                str_list = line.split()
                if str_list[0] == 'TITLE':
                    title_info += str_list
                    end_title = True
                elif end_title:
                    if 'NMR' in title_info:
                        shutil.copyfile(input_pdb, output_pdb)
                        print("cp ", input_pdb, ' -> ', output_pdb)
                    break
                line = in_f.readline()
            print(title_info)


def copy_CpHMD_pdb(total_dir, csv_path, save_pdb_dir):
    """
    This function will copy pdb files from 'total_dir' into 'save_pdb_dir' according to
    pdb id showed in csv file witch path is 'csv_path'.
    :param total_dir: String, the directory contain all proteins' molecular dynamics simulate data.
    :param csv_path: String, the path of csv file, the csv file must contain column 'PDB ID'.
    :param save_pdb_path: String, the save path of pdb files.
    :return: None
    """
    csv_df = pd.read_csv(csv_path)
    pdb_id_sr = csv_df['PDB ID'].drop_duplicates()
    if not os.path.exists(save_pdb_dir):
        os.mkdir(save_pdb_dir)
    else:
        shutil.rmtree(save_pdb_dir)
        os.mkdir(save_pdb_dir)

    for idx, pdb_id in pdb_id_sr.items():
        protein_dir = os.path.join(total_dir, '{}_output_replica'.format(pdb_id.lower()))
        input_dir = os.path.join(protein_dir, '{}_input'.format(pdb_id.lower()))
        source_pdb_path = os.path.join(input_dir, '{}.pdb'.format(pdb_id.lower()))
        with open(source_pdb_path, 'r') as f:
            line = f.readline()
            while line:
                str_list = split_pdb_line(line)
                if str_list[0] == 'ATOM' or str_list[0] == 'HETATM':
                    chain = str_list[4]
                    break
                line = f.readline()
        target_pdb_path = os.path.join(save_pdb_dir, '{}_{}.pdb'.format(pdb_id, chain))
        if os.path.exists(source_pdb_path):
            shutil.copyfile(source_pdb_path, target_pdb_path)
            print('{} -copy-> {}'.format(source_pdb_path, target_pdb_path))
        else:
            print('{} not exit'.format(source_pdb_path))


def split_pdb_line(str):
    """
    this function will split string line from pdb file, especially for string line first word is 'ATOM'.
    :param str: String, one string line form pdb file.
    :return: List[String, String ... String], a list of split string data from one pdb line.
    """
    global amino_acids_names
    str_list = str.split()
    if str_list[0] == 'ATOM' or str_list[0] == 'HETATM' and str_list[3][-3:] in amino_acids_names:
        if len(str_list) < 12:
            atom = str[0:6].split()[0]
            idx = str[6:11].split()[0]
            heavy_atom = str[11:16].split()[0]
            amino_acid = str[16:20].split()[0]
            chain = str[20:22].split()[0]
            ires_id = str[22:26].split()[0]
            x = str[26:38].split()[0]
            y = str[38:46].split()[0]
            z = str[46:54].split()[0]
            temp1 = str[54:60].split()[0]
            temp2 = str[60:66].split()[0]
            temp3 = str[66:].split()[0]
            str_list = [atom, idx, heavy_atom, amino_acid, chain, ires_id, x, y, z, temp1, temp2, temp3]
    return str_list


def stick_up_pdb_str_list(str_list):
    """
    This function will stick up str_list as a string, the string is spliced in PDB format,
    but only for list first word is 'ATOM'.
    for example:
        str_list = ['ATOM', '9', 'N', 'ILE', 'A', '2', '46.540', '32.715', '12.094', '1.00', '11.56', 'N']
     ->   string = 'ATOM      9  N   ILE A   2      46.540  32.715  12.094  1.00 11.56           N'
    :param str_list: List[String, ...], a list contain one heavy atom info, first string must be 'ATOM'.
    :return: String, pdb format string.
    """
    atom, idx, heavy_atom, amino_acid, chain, ires_id, x, y, z, temp1, temp2, temp3 = str_list
    atom = atom.ljust(6, ' ')
    idx = idx.rjust(5, ' ')
    if heavy_atom == 'SE':
        heavy_atom = heavy_atom.ljust(4, ' ').rjust(5, ' ')
    else:
        heavy_atom = heavy_atom.ljust(3, ' ').rjust(5, ' ')
    amino_acid = amino_acid.rjust(4, ' ')
    chain = chain.rjust(2, ' ')
    ires_id = ires_id.rjust(4, ' ')
    x, y, z = x.rjust(12, ' '), y.rjust(8, ' '), z.rjust(8, ' ')
    temp1, temp2, temp3 = temp1.rjust(6, ' '), temp2.rjust(6, ' '), temp3.rjust(12, ' ').ljust(14, ' ')
    return ''.join([atom, idx, heavy_atom, amino_acid, chain, ires_id, x, y, z, temp1, temp2, temp3]) + '\n'


def fix_pdb_file(source_pdb_path, target_pdb_path):
    """
    Some pdb file loss information of some group, this function will repair the loss information,
    then will drop off information of Hydrogen, at last save change information in new path.
    :param source_pdb_path: String, the path of pdb file should be repair.
    :param target_pdb_path: String, the path of new pdb file saved.
    :return: None.
    """
    # file path
    pdb_file_name = source_pdb_path.split('/')[-1].lower()
    pdb_name = pdb_file_name.split('.')[0]  # delete '.pdb'
    charmm_dir = '/data2/rymiao/propka/pka_process/charmm_script'
    temp1_pdb_path = os.path.join(charmm_dir, pdb_file_name)
    temp2_pdb_name = '{}_hsd_proa.pdb'.format(pdb_name)
    temp2_pdb_path = os.path.join(charmm_dir, temp2_pdb_name)
    charmm_inp_file = '/data2/rymiao/propka/pka_process/charmm_script/step1_pdbreader.inp'
    temp_charmm_inp_file = '/data2/rymiao/propka/pka_process/charmm_script/step1_pdbreader_temp.inp'
    fix_pdb = '/data2/rymiao/propka/pka_process/charmm_script/step1_pdbreader.pdb'
    drop_h_pdb = '/data2/rymiao/propka/pka_process/charmm_script/drop_h.pdb'

    # copy source pdb file to charmm_script dir.
    shutil.copyfile(source_pdb_path, temp1_pdb_path)
    # change 'HIS' to 'HSD', change last world to 'PROA'
    with open(temp1_pdb_path, 'r') as f1, open(temp2_pdb_path, 'w') as f2:
        contents = f1.read()
        contents = contents.replace('HIS', 'HSD')
        contents = re.sub('           [A-Z]  ', '      PROA', contents)
        f2.write(contents)

    # change 'XXXX.pdb' to temp_pdb_name
    with open(charmm_inp_file, 'r') as f1, open(temp_charmm_inp_file, 'w') as f2:
        contents = f1.read()
        contents = contents.replace('XXXX.pdb', temp2_pdb_name)
        f2.write(contents)

    # add loss information
    cmd1 = 'cd {}'.format(charmm_dir)
    cmd2 = 'charmm < {} > charmm_run_info.txt'.format(temp_charmm_inp_file.split('/')[-1])
    cmd = '&&'.join([cmd1, cmd2])
    os.system(cmd)

    # read source chain
    source_chain = ''
    with open(temp1_pdb_path, 'r') as f:
        line = f.readline()
        while (line):
            str_list = split_pdb_line(line)
            if str_list[0] == 'ATOM':
                source_chain = str_list[4]
                break
            line = f.readline()

    # drop hydrogen atom and change 'PROA' back, change chain back,
    # change 'HSD' to 'HIS', then reorder index of atoms.
    index = 1
    with open(fix_pdb, 'r') as f1, open(drop_h_pdb, 'w') as f2:
        line = f1.readline()
        while (line):
            str_list = split_pdb_line(line)
            if str_list[0] == 'ATOM':
                if str_list[2][0] != 'H':
                    str_list[1] = str(index)
                    str_list[3] = str_list[3].replace('HSD', 'HIS')
                    str_list[4] = source_chain
                    str_list[11] = str_list[2][0]
                    new_line = stick_up_pdb_str_list(str_list)
                    f2.write(new_line)
                    index = index + 1
            elif str_list[0] == 'TER':
                if len(str_list) == 5:
                    new_line = 'TER {}{}{}{}\n'.format(str(index).rjust(7, ' '), str_list[2].rjust(9, ' '),
                                                       source_chain.rjust(2, ' '), str_list[3].rjust(4, ' '))
                else:
                    new_line = 'TER\n'
                f2.write(new_line)
                break
            line = f1.readline()

    # copy drop_h.pdb to target path
    shutil.copyfile(drop_h_pdb, target_pdb_path)

    # remove temp files
    os.remove(drop_h_pdb)
    os.remove(fix_pdb)
    os.remove(temp_charmm_inp_file)
    os.remove(temp2_pdb_path)
    os.remove(temp1_pdb_path)


def fix_pdb_batch(source_dir, target_dir):
    """
    Some pdb file loss information of some group, this function will repair the loss information,
    then will drop off information of Hydrogen, at last save change information in new path.
    :param source_dir: String, the directory path contains pdb files.
    :param target_dir: String, the directory path contains fixed pdb files.
    :return: None
    """
    if not os.path.exists(target_dir):
        os.mkdir(target_dir)
    else:
        shutil.rmtree(target_dir)
        os.mkdir(target_dir)
    pdb_names = os.listdir(source_dir)
    for pdb_name in pdb_names:
        source_pdb_path = os.path.join(source_dir, pdb_name)
        target_pdb_path = os.path.join(target_dir, pdb_name)
        try:
            fix_pdb_file(source_pdb_path, target_pdb_path)
            print('{} -fixed-> {}'.format(source_pdb_path, target_pdb_path))
        except Exception as e:
            print('Fix Error: {} - {}'.format(source_pdb_path, e))


def extract_first_fixed_constant_chain(input_dir, output_dir):
    """
    This function will read pdb files from input directory, and then extract first
    constant Peptide chain saved in output directory.
    :param input_dir: String, the directory contains source pdb files.
    :param output_dir:  String, the directory to save extracted pdb files.
    :return: None
    """

    constant_num = 0
    break_num = 0
    empty_num = 0
    pdb_list = os.listdir(input_dir)
    global amino_acids_names
    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    # read pdb dir
    for pdb_name in pdb_list:
        input_pdb = os.path.join(input_dir, pdb_name)
        output_pdb = os.path.join(output_dir, pdb_name)
        # read pdb file and extract first chain composed of amino acids
        print('extract file: ', input_pdb, ' ...')
        muti_resname_ids = []
        muti_resname_dist = {}
        with open(input_pdb, 'r') as in_f:
            with open(output_pdb, 'w') as out_f:
                line = in_f.readline()
                find_chain = False
                while line:
                    str_list = split_pdb_line(line)
                    if str_list[0] == 'ATOM' or str_list[0] == 'HETATM':
                        # if str contain amino acid.
                        # print(str_list[3][-3:])
                        if str_list[3][-3:] in amino_acids_names:
                            find_chain = True
                            # ignore second amino acid info like 'BASP', 'CGGU', ...
                            if len(str_list[3]) == 4 and str_list[5] not in muti_resname_ids:
                                muti_resname_ids.append(str_list[5])
                                muti_resname_dist[str_list[5]] = str_list[3]
                                str_list[3] = str_list[3][1:]
                            elif len(str_list[3]) == 4 and str_list[3] == muti_resname_dist[str_list[5]]:
                                str_list[3] = str_list[3][1:]
                            elif len(str_list[3]) == 4:
                                line = in_f.readline()
                                continue
                            # exchange 'HETATM' to 'ATOM', 'MSE' to 'MET'
                            if str_list[0] == 'HETATM':
                                str_list[0] = 'ATOM'
                                str_list[3] = 'MET'
                                # exchange 'SE' to 'SD', 'CE' to 'C'
                                if str_list[2] == 'SE':
                                    str_list[2] = 'SD'
                                if str_list[11] == 'SE':
                                    str_list[11] = 'S'

                            new_line = stick_up_pdb_str_list(str_list)
                            out_f.write(new_line)
                    # if there is singgle chain S-S bound, write it down.
                    elif str_list[0] == 'SSBOND':
                        if str_list[3] == str_list[6]:
                            new_line = ' '.join(str_list) + '\n'
                            out_f.write(new_line)

                    if find_chain:
                        if str_list[0] == 'TER':
                            out_f.write(line)
                            break
                    line = in_f.readline()

        # read A chain pdb file, if it's residue id not constant, delete the file.
        residue_id_list = []
        with open(output_pdb, 'r') as out_f:
            line = out_f.readline()
            # print(line)
            while line:
                # print(line, end='')
                str_list = split_pdb_line(line)
                if str_list[0] == 'ATOM':
                    residue_id_list.append(int(str_list[5]))
                line = out_f.readline()
        residue_id_list = list(set(residue_id_list))
        residue_id_list.sort()
        # print(residue_id_list)
        # delete empty A chain pdb file
        if residue_id_list is None or residue_id_list == []:
            print('Delete {}, it is empty! '.format(output_pdb))
            os.remove(output_pdb)
            empty_num += 1
        else:
            list_len = len(residue_id_list)
            head_tail_range = residue_id_list[-1] - residue_id_list[0]
            if list_len < head_tail_range + 1:
                print('Delete {}, it is not constant! list_len : head_tail_range + 1 is {} : {}'
                      .format(output_pdb, list_len, head_tail_range + 1))
                os.remove(output_pdb)
                break_num += 1
            else:
                constant_num += 1

    print('count ******************')
    print('empty num: {}'.format(empty_num))
    print('break num: {}'.format(break_num))
    print('constant num: {}'.format(constant_num))


def download_pdb_from_xlsx(xlsx_filename, save_pdb_directory):
    """
    this function will read xlsx_filename, download pdb files by pdb id in xlsx_filename from website
    'https://files.rcsb.org/download/' .
    :param xlsx_filename: String, the path of the xlsx file, the xlsx file first col must content pdb id.
    :param save_pdb_directory: String, the save dir path, the downloaded pdb files will saved in the dir.
    :return: None
    """
    # read xlsx file for pdb names
    wb = load_workbook(xlsx_filename)
    sheet_name = wb.sheetnames[0]
    pdb_list = []
    idx = 0
    for row in wb[sheet_name]:
        for cell in row:
            # get rol 3 cells
            if idx == 2:
                print(cell.value)
                pdb_list.append(cell.value)
            idx = (idx + 1) % 3
    pdb_list.pop(0)

    # down pdb files
    url_prefix = 'https://files.rcsb.org/download/'
    for pdb_name in pdb_list:
        url = url_prefix + pdb_name + '.pdb'
        save_file = os.path.join(save_pdb_directory, pdb_name + '.pdb')
        r = requests.get(url)
        # print(r.content)
        print("download file: ", save_file)
        with open(save_file, "wb") as f:
            f.write(r.content)


def download_pdb_from_txt(txt_path, save_pdb_directory):
    """
    this function will read txt_path, download pdb files by pdb id in xlsx_filename from website
    'https://files.rcsb.org/download/' .
    :param txt_path: String, the path of the xlsx file, the txt file first col must content pdb id.
    :param save_pdb_directory: String, the save dir path, the downloaded pdb files will saved in the dir.
    :return: None
    """
    with open(txt_path, 'r') as f:
        content = f.read()
    str_list = content.split('\t')
    pdb_list = [word[0:4] for word in str_list]

    # down pdb files
    url_prefix = 'https://files.rcsb.org/download/'
    for pdb_name in pdb_list:
        url = url_prefix + pdb_name + '.pdb'
        save_file = os.path.join(save_pdb_directory, pdb_name + '.pdb')
        r = requests.get(url)
        # print(r.content)
        print("download file: ", save_file)
        with open(save_file, "wb") as f:
            f.write(r.content)


def download_pdb_from_csv(csv_filename, save_pdb_directory):
    """
    this function will read csv_filename, download pdb files by pdb id in csv_filename from website
    'https://files.rcsb.org/download/' .
    :param csv_filename: String, the path of the csv file, the csv file first col must content pdb id.
    :param save_pdb_directory: String, the save dir path, the downloaded pdb files will saved in the dir.
    :return: None
    """
    pdb_list = []
    idx = 0
    # 读取csv文件
    with open(csv_filename, 'r') as f:
        reader = csv.reader(f)
        result = list(reader)[1:]
    # 读取pdb id
    for line in result:
        pdb_list.append(line[0])
    # 去重
    pdb_list = list(set(pdb_list))

    # mkdir save_pdb_directory
    if not os.path.exists(save_pdb_directory):
        os.mkdir(save_pdb_directory)

    # down pdb files
    url_prefix = 'https://files.rcsb.org/download/'
    for pdb_name in pdb_list:
        url = url_prefix + pdb_name + '.pdb'
        save_file = os.path.join(save_pdb_directory, pdb_name + '.pdb')
        r = requests.get(url)
        print("download file: ", save_file)
        with open(save_file, "wb") as f:
            f.write(r.content)


def pdb_to_df(pdb_path):
    """
    this function will read pdb file, then transform  it to dataframe.
    :param pdb_path: String, the path of pdb file.
    :return pdb_df: DataFrame, the dataframe contain pdb information.
    """
    model = 1
    pdb_df = pd.DataFrame(columns=('model', 'atom', 'idx', 'heavy_atom', 'amino_acid', 'chain',
                                   'res_id', 'x', 'y', 'z', 'temp1', 'temp2', 'temp3'))
    row_list = []
    global amino_acids_names
    with open(pdb_path, 'r') as f:
        line = f.readline()
        while line:
            str_list = split_pdb_line(line)
            if str_list[0] == 'ATOM' or str_list[0] == 'HETATM' and str_list[3][-3:] in amino_acids_names:
                # pdb_df = pdb_df.append([{'model': int(model), 'atom': str_list[0], 'idx': int(str_list[1]),
                #                          'heavy_atom': str_list[2], 'amino_acid': str_list[3], 'chain': str_list[4],
                #                          'res_id': str_list[5], 'x': float(str_list[6]), 'y': float(str_list[7]),
                #                          'z': float(str_list[8]), 'temp1': str_list[9], 'temp2': str_list[10],
                #                          'temp3': str_list[11]}])
                row_list.append({'model': int(model), 'atom': str_list[0], 'idx': int(str_list[1]),
                                 'heavy_atom': str_list[2], 'amino_acid': str_list[3], 'chain': str_list[4],
                                 'res_id': str_list[5], 'x': float(str_list[6]), 'y': float(str_list[7]),
                                 'z': float(str_list[8]), 'temp1': str_list[9], 'temp2': str_list[10],
                                 'temp3': str_list[11]})
            elif str_list[0] == 'TER':
                model += 1
            line = f.readline()
    pdb_df = pdb_df.append(row_list)
    return pdb_df


def df_to_pdb_string(pdb_df):
    """
    this function will read dataframe, then transfrom it to pdb format.
    :param pdb_df: DataFrame, the dataframe contain pdb information.
    :return pdb_str: String, the pdb format of information.
    """
    pdb_str = ''
    now_model = None
    for idx, row in pdb_df.iterrows():
        model = row['model']
        if now_model is None:
            now_model = model
        elif model > now_model:
            pdb_str += 'TER\n'
            now_model = model
        str_list = [row['atom'], str(row['idx']), row['heavy_atom'], row['amino_acid'], row['chain'], row['res_id'],
                    str('%.3f' % row['x']), str('%.3f' % row['y']), str('%.3f' % row['z']), row['temp1'], row['temp2'],
                    row['temp3']]
        line = stick_up_pdb_str_list(str_list)
        pdb_str += line
    pdb_str += 'TER\n'
    return pdb_str


def get_SSBOND_content(pdb_path):
    """
    this function will read pdb file, then get single chain S-S bond information, and return it.
    :param pdb_path: String, the path of pdb file.
    :return content: String, the content of single S-S bond information.
    """
    content = ''
    with open(pdb_path, 'r') as f:
        line = f.readline()
        while (line):
            str_list = split_pdb_line(line)
            if str_list[0] == 'SSBOND':
                if str_list[3] == str_list[6]:
                    new_line = ' '.join(str_list)
                    content += new_line + '\n'
            elif str_list[0] == 'ATOM' or str_list[0] == 'HETATM':
                break
            line = f.readline()
    return content


def extract_single_chain(csv_file, pdb_dir, single_chain_pdb_dir):
    """
    this function will extract single chain info from pdb_dir to singe_chain_pdb_dir according to csv file.
    :param csv_file: String, path of csv file, the csv file must contain columns ['PDB ID' 'Chain']
    :param pdb_dir: String, path of directory contain sources pdb files.
    :param single_chain_pdb_dir: String, path of directory should save extracted pdb files.
    :return: None.
    """
    csv_df = pd.read_csv(csv_file)
    csv_df = csv_df.loc[:, ['PDB ID', 'Chain']].drop_duplicates()
    csv_df = csv_df.sort_values(by=['PDB ID', 'Chain'])
    # create output dir
    if os.path.exists(single_chain_pdb_dir):
        shutil.rmtree(single_chain_pdb_dir)
        os.mkdir(single_chain_pdb_dir)
    else:
        os.mkdir(single_chain_pdb_dir)

    for idx, row in csv_df.iterrows():
        pdb_path = os.path.join(pdb_dir, '{}.pdb'.format(row['PDB ID']))
        if not os.path.exists(pdb_path):
            print('{} not exist!'.format(pdb_path))
            continue
        SSBOND_content = get_SSBOND_content(pdb_path)
        pdb_df = pdb_to_df(pdb_path)
        for idx2, row2 in csv_df.loc[csv_df['PDB ID'] == row['PDB ID']].iterrows():
            chain_pdb_df = pdb_df.loc[pdb_df['chain'] == row2['Chain']]
            min_model = chain_pdb_df['model'].min()
            single_chain_pdb_df = chain_pdb_df.loc[chain_pdb_df['model'] == min_model]
            single_chain_pdb_str = df_to_pdb_string(single_chain_pdb_df)
            save_name = '{}_{}.pdb'.format(row['PDB ID'][0:4], row2['Chain'])
            save_path = os.path.join(single_chain_pdb_dir, save_name)
            print('save extract pdb file: {}'.format(save_path))
            with open(save_path, 'w') as f:
                union_content = SSBOND_content + single_chain_pdb_str
                f.write(union_content)


def crop_pdb_ires_v1(csv_filename, pdb_directory, crop_pdb_directory, wrong_ires_info_file, center_coors_dir):
    """
    this function will read csv_filename and pdb files in pdb_directory, according to titratable residue center
    to crop heavy atoms info from original pdb file into new pdb files. And create a csv file to save titratable
     residue center coordinates.

    :param csv_filename: String, the path of the csv file, the csv file must contain 'pdb id', 'Res Name', 'Chain',
    'Res ID', 'Expt. pKa' information.
    :param pdb_directory: String, the path of directory contain pdb files labeled in csv file.
    :param crop_pdb_directory: String, the path of directory saved crop pdb files.
    :param wrong_ires_info_file: String, the path of file contain error information of titratable residue.
    :param center_coors_dir: String, the directory to center coors csv files.
    :return: None
    """
    # read csv file
    csv_df = pd.read_csv(csv_filename)
    # change ['PDB ID', 'Res Name', 'Chain', 'Res ID', 'Expt. pKa', 'Expt. Uncertainty',
    # '%SASA', 'Expt. method', 'Expt. salt conc', 'Expt. pH', 'Expt. temp', 'Reference']
    # or other formats
    # to ['PDB ID', 'Chain', 'Res ID', 'Res Name', 'pKa']
    column_order = ['PDB ID', 'Chain', 'Res ID', 'Res Name', 'pKa']
    csv_df = csv_df.rename(columns={'Expt pKa': 'pKa'})  # change 'Expt. pKa' to 'pKa'
    csv_df = csv_df.loc[:, ['PDB ID', 'Chain', 'Res ID', 'Res Name', 'pKa']]  # screen columns,  order columns

    # save new csv in '.temp.csv'
    temp_path = '.temp.csv'
    csv_df.to_csv(temp_path, index=False)

    # read new csv file data
    with open(temp_path, 'r') as f:
        reader = csv.reader(f)
        result = list(reader)[1:]
    os.remove(temp_path)
    pdb_info = {}

    # get info from csv
    '''
        pdb_info = {
            'PDB ID0'
                'Chain0': {
                    'ResID0': {
                        'ResName': '',
                        'pKa':'',
                        'SASA':''  
                    },
                    'ResID1':{
                        'ResName': '',
                        'pKa':'',
                        'SASA':'' 
                    }
                    ...
                }
                'Chain1': {
                    'ResID0': {
                        'ResName': '',
                        'pKa':'',
                        'SASA':''  
                    },
                    'ResID1':{
                        'ResName': '',
                        'pKa':'',
                        'SASA':'' 
                    } 
                    ...
                }
                ...
            },
            'PDB ID1':{
                'Chain0': {
                    'ResID0': {
                        'ResName': '',
                        'pKa':'',
                        'SASA':''  
                    },
                    'ResID1':{
                        'ResName': '',
                        'pKa':'',
                        'SASA':'' 
                    }
                    ...
                ...
            },
            ...
        }
    '''
    for line in result:
        if line[0] not in pdb_info.keys():
            pdb_info[line[0]] = {}
        if line[1] not in pdb_info[line[0]].keys():
            pdb_info[line[0]][line[1]] = {}
        if line[2] not in pdb_info[line[0]][line[1]].keys():
            pdb_info[line[0]][line[1]][line[2]] = {}
        pdb_info[line[0]][line[1]][line[2]]['ResName'] = line[3]
        pdb_info[line[0]][line[1]][line[2]]['pKa'] = line[4]
        # pdb_info[line[0]][line[1]][line[2]]['SASA'] = line[6]

    # crop pdb file with ires center
    crop_l = 21
    crop_w = 21
    crop_h = 21
    count_wrong_ires = 0
    count_sucess_ires = 0
    wrong_ires_list = []

    # make crop directory
    if not os.path.exists(crop_pdb_directory):
        os.mkdir(crop_pdb_directory)
    else:
        shutil.rmtree(crop_pdb_directory)
        os.mkdir(crop_pdb_directory)

    file_name = []
    center_x = []
    center_y = []
    center_z = []
    for pdb_id in pdb_info.keys():
        # dump over wrong file
        error_files = []
        if pdb_id in error_files:
            continue

        # extract chain of atoms info from pdb file
        for chain in pdb_info[pdb_id].keys():
            # read pdb file
            pdb_path = os.path.join(pdb_directory, '{}_{}.pdb'.format(pdb_id, chain))
            try:
                with open(pdb_path, 'r') as in_f:
                    print('read file :', pdb_path)
                    chain_lines = []  # one chain of atoms info
                    line = in_f.readline()
                    has_find_chain = False
                    while line:
                        str_list = split_pdb_line(line)
                        if str_list[0] == 'ATOM':
                            if str_list[4] == chain:
                                chain_lines.append(line)
                                has_find_chain = True

                        if has_find_chain:
                            if str_list[0] == 'TER':
                                break
                        line = in_f.readline()

                    for ires_id in pdb_info[pdb_id][chain].keys():
                        # Calculate the central coordinates of the titration point
                        ires_name = pdb_info[pdb_id][chain][ires_id]['ResName'];
                        ires_center = [0, 0, 0]
                        ires_name_list = []
                        p1_coor = []
                        p2_coor = []
                        if ires_name == 'ASP':
                            ires_name_list = ['ASP', 'AASP']
                            p1_name = 'OD1'
                            p2_name = 'OD2'
                        elif ires_name == 'GLU':
                            ires_name_list = ['GLU', 'AGLU']
                            p1_name = 'OE1'
                            p2_name = 'OE2'
                        elif ires_name == 'LYS':
                            ires_name_list = ['LYS', 'ALYS']
                            p1_name = 'NZ'
                            p2_name = 'NZ'
                        elif ires_name == 'HIS':
                            ires_name_list = ['HIS', 'AHIS']
                            p1_name = 'ND1'
                            p2_name = 'NE2'
                        elif ires_name == 'CYS':
                            ires_name_list = ['CYS', 'ACYS']
                            p1_name = 'SG'
                            p2_name = 'SG'
                        else:
                            continue

                        for chain_line in chain_lines:
                            str_list = split_pdb_line(chain_line)
                            if str_list[5] == ires_id and str_list[3] in ires_name_list:
                                # atom coordinate
                                coor = np.array([float(str_list[6]), float(str_list[7]), float(str_list[8])])
                                if str_list[2] == p1_name:
                                    p1_coor = coor
                                if str_list[2] == p2_name:
                                    p2_coor = coor
                                    break
                        try:
                            ires_center = (p1_coor + p2_coor) / 2
                        except:
                            print('error :', 'ires_name = ', ires_name, ' ,ires_id = ', ires_id)
                            print('p1_name:', p1_name, 'p2_name:', p2_name, 'ires_name_list:', ires_name_list)
                            print('p1_coor: ', p1_coor)
                            print('p2_coor: ', p2_coor)
                            for chain_line in chain_lines:
                                str_list = split_pdb_line(chain_line)
                                if str_list[5] == ires_id and str_list[3] in ires_name_list:
                                    print(chain_line, str_list)

                            # exit(-1)
                            count_wrong_ires += 1
                            wrong_ires_list.append(pdb_id + ' ' + chain + ' ' + ires_id + ' ' + ires_name + '\n')
                            continue

                        crop_lines = []
                        crop_pdb_path = os.path.join(crop_pdb_directory,
                                                     pdb_id + '_' + chain + '_' + ires_id + '_' + ires_name + '.pdb')
                        amino_acid_ids = []
                        with open(crop_pdb_path, 'w') as out_f:
                            for chain_line in chain_lines:
                                str_list = split_pdb_line(chain_line)
                                # get amino_acid_ids
                                try:
                                    amino_acid_id = int(str_list[5])
                                except ValueError:
                                    continue
                                # atom coordinate
                                coor = np.array([float(str_list[6]), float(str_list[7]), float(str_list[8])])
                                # the vector of ires center point to atom point
                                vec = coor - ires_center
                                # read atom amino aci name
                                amino_aci_name = str_list[3]
                                if len(amino_aci_name) == 4:
                                    if amino_aci_name[0] == 'B':
                                        continue
                                if abs(vec[0]) <= crop_l / 2:
                                    if abs(vec[1]) <= crop_w / 2:
                                        if abs(vec[2]) <= crop_h / 2:
                                            if amino_acid_id not in amino_acid_ids:
                                                amino_acid_ids.append(amino_acid_id)
                                            x = round(vec[0], 3)
                                            y = round(vec[1], 3)
                                            z = round(vec[2], 3)
                                            x_str = str(x).rjust(8, ' ')
                                            y_str = str(y).rjust(8, ' ')
                                            z_str = str(z).rjust(8, ' ')
                                            new_line = chain_line[0:30] + x_str + y_str + z_str + chain_line[54:]
                                            out_f.write(new_line)
                            out_f.write('TER')
                            count_sucess_ires += 1
                        new_res_id = amino_acid_ids.index(int(ires_id))
                        new_crop_pdb_name = '{}_{}_{}_{}_{}.pdb'.format(pdb_id, chain, ires_id, ires_name,
                                                                        str(new_res_id))
                        new_crop_pdb_path = os.path.join(crop_pdb_directory, new_crop_pdb_name)
                        os.rename(crop_pdb_path, new_crop_pdb_path)
                        if new_crop_pdb_name.split('.')[0] == '2WBF_X_694_GLU_6_X':
                            print(new_res_id)
                            exit(0)
                        file_name.append(new_crop_pdb_name.split('.')[0])
                        center_x.append(ires_center[0])
                        center_y.append(ires_center[1])
                        center_z.append(ires_center[2])
                        print('create file :', new_crop_pdb_path)

            except Exception as e:
                print(e)
                continue

    # save titrable residue center coordinates in csv file
    if not os.path.exists(center_coors_dir):
        os.mkdir(center_coors_dir)
    center_coor_pd = pd.DataFrame(data={'file_name': file_name, 'x': center_x, 'y': center_y, 'z': center_z})
    center_coor_csv_name = '{}_center_coors.csv'.format(csv_filename.split('/')[-1].split('.')[0])
    center_coor_csv_path = os.path.join(center_coors_dir, center_coor_csv_name)
    print(center_coors_dir)
    print(center_coor_csv_name)
    center_coor_pd.to_csv(center_coor_csv_path, index=False)
    print('center coors file saved in {}'.format(center_coor_csv_path))

    # write wrong info to file
    with open(wrong_ires_info_file, 'w') as info_f:
        info_f.write("pdbID chain iresID iresName\n")
        for each in wrong_ires_list:
            info_f.write(each)

    print('all done!')
    print('count wrong ires:', count_wrong_ires)
    print('count sucess ires', count_sucess_ires)


def crop_pdb_ires_v2(csv_filename, pdb_directory, crop_pdb_directory, crop_radii=10, crop_type='cube'):
    """
    this function will read csv_filename and pdb files in pdb_directory, according to titratable residue center
    to crop heavy atoms info from original pdb file into new pdb files.

    :param csv_filename: String, the path of the csv file, the csv file must contain 'pdb id', 'Res Name', 'Chain',
    'Res ID', 'Expt. pKa' information.
    :param pdb_directory: String, the path of directory contain pdb files labeled in csv file.
    :param crop_pdb_directory: String, the path of directory saved crop pdb files.
    :return: None
    """
    # 1. read csv file, transform it to DataFrame
    pka_info = pd.read_csv(csv_filename)
    # pka_info = pka_info.T
    # new_pka_info = pka_info['PDB ID'] + pka_info['Chain'] + pka_info['Res ID'] \
    # + pka_info['Res Name'] + pka_info['Expt. pKa']

    new_pka_info = pka_info.loc[:, ['PDB ID', 'Chain', 'Res ID', 'Res Name', 'Expt. pKa']]

    print(new_pka_info)
    # unfinished
    pass


def protonated_and_charged(pdb_dir, mol2_dir):
    """
    This function will use UCSF Chimera with Amber ff14SB for standard residues to transform pdb files to mol2 files.
    :param pdb_dir: input directory contain pdb files.
    :param mol2_dir: output directory contain mol2 files.
    :return: None.
    """
    if not os.path.exists(mol2_dir):
        os.mkdir(mol2_dir)
    pdb_list = os.listdir(pdb_dir)

    tmp_path = os.path.join(mol2_dir, 'tmp.mol2')
    for pdb_name in pdb_list:
        mol2_name = pdb_name.replace('.pdb', '.mol2')
        pdb_path = os.path.join(pdb_dir, pdb_name)
        mol2_path = os.path.join(mol2_dir, mol2_name)
        # prepare cmd
        cmd_content = 'open {} \n addh \n addcharge \n write format mol2 0 {} \n stop'.format(pdb_path, tmp_path)
        with open('tmp.sh', 'w') as f:
            f.write(cmd_content)
        # use UCSF Chimera to protonated and charged
        cmd1 = 'cat tmp.sh | chimera --nogui'.format(pdb_path, tmp_path)
        # Do not use TIP3P atom types, pybel cannot read them
        cmd2 = 'sed \'s/H\.t3p/H    /\' {} | sed \'s/O\.t3p/O\.3  /\' > {}'.format(tmp_path, mol2_path)
        cmd = '{} && {}'.format(cmd1, cmd2)
        os.system(cmd)
        print('*' * 100)
        print(cmd1)
        print(cmd2)
        os.system('cat tmp.sh')
        print('{} -UCSF Chimera-> {}'.format(pdb_path, mol2_path))
        print('*' * 100)
    # delete tmp.mol2 file
    cmd3 = 'rm {} && rm tmp.sh'.format(tmp_path)
    print(cmd3)
    os.system(cmd3)

